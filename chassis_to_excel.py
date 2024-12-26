#!/usr/bin/python
# -*- coding: utf-8 -*-

import glob
import datetime

#para trabajar con Excel
import openpyxl
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill


#############################################################################################################
#Command "show chassis hardware clei-models" must be dumped into excel files from routers
#
#############################################################################################################



my_mx_dict = {
        "name": [],
        "chassis" : [],
    "cb_0": [],
    "cb_1": [],
    "cb_2": [],
    "re_0": [],
    "re_1": [],
    "fpc_0": [ "", "", "" ],
    "fpc_1": [ "", "", "" ],
    "fpc_2": [ "", "", "" ],
    "fpc_3": [ "", "", "" ],
    "fpc_4": [ "", "", "" ],
    "fpc_5": [ "", "", "" ],
    "fpc_6": [ "", "", "" ],
    "fpc_7": [ "", "", "" ],
    "fpc_8": [ "", "", "" ],
    "fpc_9": [ "", "", "" ],
    "fpc_10": [ "", "", "" ],
    "fpc_11": [ "", "", "" ]
}

#############################################################################################################
#Def get_router_name
#  This funtion gets the name of the router
#############################################################################################################

def get_router_name (def_routerFile):
    varTmp1 = def_routerFile.split('/')
    #print ('varTmp1 vale ', varTmp1)
    def_routerName = varTmp1[8].replace('.show-chassis-hw.xml','')
    return (def_routerName)


#############################################################################################################
#Def draw_router_hw
#  This funtion dumps HW file into dictionary
#############################################################################################################

def draw_router_hw (def_hw_inventory, def_mx_dict):
    def_file_inventory = open(def_hw_inventory, 'r')
    #print ('def_file_inventory vale: ', def_file_inventory)

    #def_mx_dict = my_mx_dict

    for line in def_file_inventory:
        if (line.find('Midplane') != -1):
            varTmp1 = line.split()
            #def_chassis = varTmp1[2]
            print ('varTmp1 vale ', varTmp1)
            def_mx_dict["chassis"] = varTmp1[5]
            #print ('chassis is ', def_chassis)
            continue
        if (line.find('Routing Engine 0') != -1):
            varTmp1 = line.split()
            #def_re0 = varTmp1[7]
            def_mx_dict["re_0"] = varTmp1[7]
            #print ('RE0 is ', def_re0)
            continue
        if (line.find('Routing Engine 1') != -1):
            varTmp1 = line.split()
            #def_re1 = varTmp1[7]
            def_mx_dict["re_1"] = varTmp1[7]
            #print ('RE1 is ', def_re1)
            continue
        if (line.find('CB 0') != -1):
            varTmp1 = line.split()
            #def_cb0 = varTmp1[7]
            def_mx_dict["cb_0"] = varTmp1[6]
            #print ('CB0 is ', def_cb0)
            continue
        if (line.find('CB 1') != -1):
            varTmp1 = line.split()
            #def_cb1 = varTmp1[7]
            def_mx_dict["cb_1"] = varTmp1[6]
            #print ('CB1 is ', def_cb1)
            continue
        if (line.find('CB 2') != -1):
            varTmp1 = line.split()
            #def_cb1 = varTmp1[7]
            def_mx_dict["cb_2"] = varTmp1[6]
            #print ('CB1 is ', def_cb1)
            continue
        if (line.find('FPC') != -1):
            print('line es: ',line)
            varTmp1 = line.split()
            def_slot = ''.join(['fpc_',varTmp1[1]])
            if (line.find('BUILTIN') != -1):
                def_mx_dict[def_slot][0] = 'BUILTIN'
            elif (line.find('MS-MPC') != -1):
                def_mx_dict[def_slot][0] = varTmp1[6]
                def_mx_dict[def_slot][1] = 'NA'
                def_mx_dict[def_slot][2] = 'NA'
            elif (line.find('MPC7E') != -1):
                def_mx_dict[def_slot][0] = varTmp1[6]
                def_mx_dict[def_slot][1] = 'NA'
                def_mx_dict[def_slot][2] = 'NA'
            else:
                def_mx_dict[def_slot][0] = varTmp1[6]
            continue
        if (line.find('MIC') != -1):
            varTmp1 = line.split()
            def_mic = int(varTmp1[1]) + 1
            #def_slot comes from previous "FPC" loop
            #print ('def_mic es: ', def_mic)
            def_mx_dict[def_slot][def_mic] = varTmp1[6]
            continue
    """
    print('name es: ', def_mx_dict["name"])
    print('chassis es: ', def_mx_dict["chassis"])
    print('CB0 es: ', def_mx_dict["cb_0"])
    print('CB1 es: ', def_mx_dict["cb_1"])
    print('CB2 es: ', def_mx_dict["cb_2"])
    print('RE0 es: ', def_mx_dict["re_0"])
    print('RE1 es: ', def_mx_dict["re_1"])
    print('FPC0 es: ', def_mx_dict["fpc_0"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print('FPC1 es: ', def_mx_dict["fpc_1"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print('FPC2 es: ', def_mx_dict["fpc_2"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print('FPC3 es: ', def_mx_dict["fpc_3"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print('FPC4 es: ', def_mx_dict["fpc_4"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print('FPC5 es: ', def_mx_dict["fpc_5"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print('FPC7 es: ', def_mx_dict["fpc_7"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print('FPC8 es: ', def_mx_dict["fpc_8"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print('FPC9 es: ', def_mx_dict["fpc_9"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print('FPC10 es: ', def_mx_dict["fpc_10"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print('FPC11 es: ', def_mx_dict["fpc_11"][0])
    print('  MIC0 es: ', def_mx_dict["fpc_0"][1])
    print('  MIC1 es: ', def_mx_dict["fpc_0"][2])
    print ('####################')
    """

    return (def_mx_dict)


#############################################################################################################
#Def draw_layout
#  This funtion draws the layout
#############################################################################################################

def draw_layout (def_mx_dict):

    varTmp1 = def_mx_dict["chassis"].split('-')
    for item in varTmp1:
        if (item.find('MX') !=-1):
            def_model_mx = item
            break

    referenceSheet = wb.get_sheet_by_name(def_model_mx)

    current_sheet = wb.copy_worksheet(referenceSheet)
    current_sheet.title = def_mx_dict["name"]
    current_sheet['C3'] = def_mx_dict["name"]

    bold_font_white = Font(bold=True, color="806000", size=12)
    #center_aligned_text = Alignment(horizontal="center")
    yellow_mpc_background = PatternFill(patternType="solid", fgColor="FBE7A3")
    yellow_mic_background = PatternFill(patternType="solid", fgColor="FFF2CC")

    if (def_model_mx == 'MX960'):
        if (def_mx_dict["cb_0"]):
            #SCB0
            current_sheet['I5'] = def_mx_dict["cb_0"]
        if (def_mx_dict["cb_1"]):
            #SCB1
            current_sheet['J5'] = def_mx_dict["cb_1"]
        if (def_mx_dict["cb_2"]):
            #SCB2
            current_sheet['K5'] = def_mx_dict["cb_2"]
        if (def_mx_dict["re_0"]):
            #RE0
            current_sheet['I7'] = def_mx_dict["re_0"]
        if (def_mx_dict["re_1"]):
            #RE1
            current_sheet['J7'] = def_mx_dict["re_1"]

        if (def_mx_dict["fpc_0"][0]):
            #FPC0
            if (def_mx_dict["fpc_0"][1]):
                current_sheet['C5'] = def_mx_dict["fpc_0"][1]
                current_sheet['C5'].font = bold_font_white
                current_sheet['C5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_0"][2]):
                current_sheet['C6'] = def_mx_dict["fpc_0"][2]
                current_sheet['C6'].font = bold_font_white
                current_sheet['C6'].fill = yellow_mic_background
            current_sheet['C7'] = def_mx_dict["fpc_0"][0]
            current_sheet['C7'].font = bold_font_white
            current_sheet['C7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_1"][0]):
            #FPC1
            if (def_mx_dict["fpc_1"][1]):
                current_sheet['D5'] = def_mx_dict["fpc_1"][1]
                current_sheet['D5'].font = bold_font_white
                current_sheet['D5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_1"][2]):
                current_sheet['D6'] = def_mx_dict["fpc_1"][2]
                current_sheet['D6'].font = bold_font_white
                current_sheet['D6'].fill = yellow_mic_background
            current_sheet['D7'] = def_mx_dict["fpc_1"][0]
            current_sheet['D7'].font = bold_font_white
            current_sheet['D7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_2"][0]):
            #FPC2
            if (def_mx_dict["fpc_2"][1]):
                current_sheet['E5'] = def_mx_dict["fpc_2"][1]
                current_sheet['E5'].font = bold_font_white
                current_sheet['E5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_2"][2]):
                current_sheet['E6'] = def_mx_dict["fpc_2"][2]
                current_sheet['E6'].font = bold_font_white
                current_sheet['E6'].fill = yellow_mic_background
            current_sheet['E7'] = def_mx_dict["fpc_2"][0]
            current_sheet['E7'].font = bold_font_white
            current_sheet['E7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_3"][0]):
            #FPC3
            if (def_mx_dict["fpc_3"][1]):
                current_sheet['F5'] = def_mx_dict["fpc_3"][1]
                current_sheet['F5'].font = bold_font_white
                current_sheet['F5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_3"][2]):
                current_sheet['F6'] = def_mx_dict["fpc_3"][2]
                current_sheet['F6'].font = bold_font_white
                current_sheet['F6'].fill = yellow_mic_background
            current_sheet['F7'] = def_mx_dict["fpc_3"][0]
            current_sheet['F7'].font = bold_font_white
            current_sheet['F7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_4"][0]):
            #FPC4
            if (def_mx_dict["fpc_4"][1]):
                current_sheet['G5'] = def_mx_dict["fpc_4"][1]
                current_sheet['G5'].font = bold_font_white
                current_sheet['G5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_4"][2]):
                current_sheet['G6'] = def_mx_dict["fpc_4"][2]
                current_sheet['G6'].font = bold_font_white
                current_sheet['G6'].fill = yellow_mic_background
            current_sheet['G7'] = def_mx_dict["fpc_4"][0]
            current_sheet['G7'].font = bold_font_white
            current_sheet['G7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_5"][0]):
            #FPC5
            if (def_mx_dict["fpc_5"][1]):
                current_sheet['H5'] = def_mx_dict["fpc_5"][1]
                current_sheet['H5'].font = bold_font_white
                current_sheet['H5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_5"][2]):
                current_sheet['H6'] = def_mx_dict["fpc_5"][2]
                current_sheet['H6'].font = bold_font_white
                current_sheet['H6'].fill = yellow_mic_background
            current_sheet['H7'] = def_mx_dict["fpc_5"][0]
            current_sheet['H7'].font = bold_font_white
            current_sheet['H7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_7"][0]):
            #FPC7
            if (def_mx_dict["fpc_7"][1]):
                current_sheet['L5'] = def_mx_dict["fpc_7"][1]
                current_sheet['L5'].font = bold_font_white
                current_sheet['L5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_7"][2]):
                current_sheet['L6'] = def_mx_dict["fpc_7"][2]
                current_sheet['L6'].font = bold_font_white
                current_sheet['L6'].fill = yellow_mic_background
            current_sheet['L7'] = def_mx_dict["fpc_7"][0]
            current_sheet['L7'].font = bold_font_white
            current_sheet['L7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_8"][0]):
            #FPC8
            if (def_mx_dict["fpc_8"][1]):
                current_sheet['M5'] = def_mx_dict["fpc_8"][1]
                current_sheet['M5'].font = bold_font_white
                current_sheet['M5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_8"][2]):
                current_sheet['M6'] = def_mx_dict["fpc_8"][2]
                current_sheet['M6'].font = bold_font_white
                current_sheet['M6'].fill = yellow_mic_background
            current_sheet['M7'] = def_mx_dict["fpc_8"][0]
            current_sheet['M7'].font = bold_font_white
            current_sheet['M7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_9"][0]):
            #FPC9
            if (def_mx_dict["fpc_9"][1]):
                current_sheet['N5'] = def_mx_dict["fpc_9"][1]
                current_sheet['N5'].font = bold_font_white
                current_sheet['N5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_9"][2]):
                current_sheet['N6'] = def_mx_dict["fpc_9"][2]
                current_sheet['N6'].font = bold_font_white
                current_sheet['N6'].fill = yellow_mic_background
            current_sheet['N7'] = def_mx_dict["fpc_9"][0]
            current_sheet['N7'].font = bold_font_white
            current_sheet['N7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_10"][0]):
            #FPC10
            if (def_mx_dict["fpc_10"][1]):
                current_sheet['O5'] = def_mx_dict["fpc_10"][1]
                current_sheet['O5'].font = bold_font_white
                current_sheet['O5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_10"][2]):
                current_sheet['O6'] = def_mx_dict["fpc_10"][2]
                current_sheet['O6'].font = bold_font_white
                current_sheet['O6'].fill = yellow_mic_background
            current_sheet['O7'] = def_mx_dict["fpc_10"][0]
            current_sheet['O7'].font = bold_font_white
            current_sheet['O7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_11"][0]):
            #FPC11
            if (def_mx_dict["fpc_11"][1]):
                current_sheet['P5'] = def_mx_dict["fpc_11"][1]
                current_sheet['P5'].font = bold_font_white
                current_sheet['P5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_11"][2]):
                current_sheet['P6'] = def_mx_dict["fpc_11"][2]
                current_sheet['P6'].font = bold_font_white
                current_sheet['P6'].fill = yellow_mic_background
            current_sheet['P7'] = def_mx_dict["fpc_11"][0]
            current_sheet['P7'].font = bold_font_white
            current_sheet['P7'].fill = yellow_mpc_background

    if (def_model_mx == 'MX480'):
        if (def_mx_dict["cb_0"]):
            #SCB0
            current_sheet['C12'] = def_mx_dict["cb_0"]
        if (def_mx_dict["cb_1"]):
            #SCB1
            current_sheet['C11'] = def_mx_dict["cb_1"]
        if (def_mx_dict["re_0"]):
            #RE0
            current_sheet['D12'] = def_mx_dict["re_0"]
        if (def_mx_dict["re_1"]):
            #RE1
            current_sheet['D11'] = def_mx_dict["re_1"]

        if (def_mx_dict["fpc_0"][0]):
            #FPC0
            if (def_mx_dict["fpc_0"][2]):
                current_sheet['E10'] = def_mx_dict["fpc_0"][2]
                current_sheet['E10'].font = bold_font_white
                current_sheet['E10'].fill = yellow_mic_background
            if (def_mx_dict["fpc_0"][1]):
                current_sheet['D10'] = def_mx_dict["fpc_0"][1]
                current_sheet['D10'].font = bold_font_white
                current_sheet['D10'].fill = yellow_mic_background
            current_sheet['C10'] = def_mx_dict["fpc_0"][0]
            current_sheet['C10'].font = bold_font_white
            current_sheet['C10'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_1"][0]):
            #FPC1
            if (def_mx_dict["fpc_1"][2]):
                current_sheet['E9'] = def_mx_dict["fpc_1"][2]
                current_sheet['E9'].font = bold_font_white
                current_sheet['E9'].fill = yellow_mic_background
            if (def_mx_dict["fpc_1"][1]):
                current_sheet['D9'] = def_mx_dict["fpc_1"][1]
                current_sheet['D9'].font = bold_font_white
                current_sheet['D9'].fill = yellow_mic_background
            current_sheet['C9'] = def_mx_dict["fpc_1"][0]
            current_sheet['C9'].font = bold_font_white
            current_sheet['C9'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_2"][0]):
            #FPC2
            if (def_mx_dict["fpc_2"][2]):
                current_sheet['E8'] = def_mx_dict["fpc_2"][2]
                current_sheet['E8'].font = bold_font_white
                current_sheet['E8'].fill = yellow_mic_background
            if (def_mx_dict["fpc_2"][1]):
                current_sheet['D8'] = def_mx_dict["fpc_2"][1]
                current_sheet['D8'].font = bold_font_white
                current_sheet['D8'].fill = yellow_mic_background
            current_sheet['C8'] = def_mx_dict["fpc_2"][0]
            current_sheet['C8'].font = bold_font_white
            current_sheet['C8'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_3"][0]):
            #FPC3
            if (def_mx_dict["fpc_3"][2]):
                current_sheet['E7'] = def_mx_dict["fpc_3"][2]
                current_sheet['E7'].font = bold_font_white
                current_sheet['E7'].fill = yellow_mic_background
            if (def_mx_dict["fpc_3"][1]):
                current_sheet['D7'] = def_mx_dict["fpc_3"][1]
                current_sheet['D7'].font = bold_font_white
                current_sheet['D7'].fill = yellow_mic_background
            current_sheet['C7'] = def_mx_dict["fpc_3"][0]
            current_sheet['C7'].font = bold_font_white
            current_sheet['C7'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_4"][0]):
            #FPC4
            if (def_mx_dict["fpc_4"][2]):
                current_sheet['E6'] = def_mx_dict["fpc_4"][2]
                current_sheet['E6'].font = bold_font_white
                current_sheet['E6'].fill = yellow_mic_background
            if (def_mx_dict["fpc_4"][1]):
                current_sheet['D6'] = def_mx_dict["fpc_4"][1]
                current_sheet['D6'].font = bold_font_white
                current_sheet['D6'].fill = yellow_mic_background
            current_sheet['C6'] = def_mx_dict["fpc_4"][0]
            current_sheet['C6'].font = bold_font_white
            current_sheet['C6'].fill = yellow_mpc_background

        if (def_mx_dict["fpc_5"][0]):
            #FPC5
            if (def_mx_dict["fpc_5"][2]):
                current_sheet['E5'] = def_mx_dict["fpc_5"][2]
                current_sheet['E5'].font = bold_font_white
                current_sheet['E5'].fill = yellow_mic_background
            if (def_mx_dict["fpc_5"][1]):
                current_sheet['D5'] = def_mx_dict["fpc_5"][1]
                current_sheet['D5'].font = bold_font_white
                current_sheet['D5'].fill = yellow_mic_background
            current_sheet['C5'] = def_mx_dict["fpc_5"][0]
            current_sheet['C5'].font = bold_font_white
            current_sheet['C5'].fill = yellow_mpc_background

    if (def_model_mx == 'MX104'):
        if (def_mx_dict["re_0"]):
            #RE0
            current_sheet['D9'] = def_mx_dict["re_0"]
        if (def_mx_dict["re_1"]):
            #RE1
            current_sheet['D8'] = def_mx_dict["re_1"]

        if (def_mx_dict["fpc_0"][0]):
            #FPC0
            if (def_mx_dict["fpc_0"][2]):
                current_sheet['E7'] = def_mx_dict["fpc_0"][2]
                current_sheet['E7'].font = bold_font_white
                current_sheet['E7'].fill = yellow_mic_background
            if (def_mx_dict["fpc_0"][1]):
                current_sheet['D7'] = def_mx_dict["fpc_0"][1]
                current_sheet['D7'].font = bold_font_white
                current_sheet['D7'].fill = yellow_mic_background
        if (def_mx_dict["fpc_1"][0]):
            #FPC1
            if (def_mx_dict["fpc_1"][2]):
                current_sheet['E6'] = def_mx_dict["fpc_1"][2]
                current_sheet['E6'].font = bold_font_white
                current_sheet['E6'].fill = yellow_mic_background
            if (def_mx_dict["fpc_1"][1]):
                current_sheet['D6'] = def_mx_dict["fpc_1"][1]
                current_sheet['D6'].font = bold_font_white
                current_sheet['D6'].fill = yellow_mic_background

#############################################################################################################
# Main Function
#
#############################################################################################################


#Open Excel file
pathFiles = '/Users/jomih/Documents/20.Juniper/100.Scripts/chassis_to_excel/excel_reference/'

varTmp1 = ''.join([pathFiles,'/ExcelReference.xlsx'])
wb = openpyxl.load_workbook(varTmp1)

#Open HW files
pathFiles = '/Users/jomih/Documents/20.Juniper/100.Scripts/chassis_to_excel/hw_files/'
varTmp1 = ''.join([pathFiles,'/*show-chassis*.xml'])
listFiles = glob.glob(varTmp1)

for routerFile in listFiles:

    mx_dict = {
        "name": [],
        "chassis" : [],
        "cb_0": [],
        "cb_1": [],
        "cb_2": [],
        "re_0": [],
        "re_1": [],
        "fpc_0": [ "", "", "" ],
        "fpc_1": [ "", "", "" ],
        "fpc_2": [ "", "", "" ],
        "fpc_3": [ "", "", "" ],
        "fpc_4": [ "", "", "" ],
        "fpc_5": [ "", "", "" ],
        "fpc_6": [ "", "", "" ],
        "fpc_7": [ "", "", "" ],
        "fpc_8": [ "", "", "" ],
        "fpc_9": [ "", "", "" ],
        "fpc_10": [ "", "", "" ],
        "fpc_11": [ "", "", "" ]
    }

    #dump router HW
    mx_dict = draw_router_hw(routerFile, mx_dict)

    #Get router Name
    routerName = get_router_name(routerFile)
    mx_dict["name"] = routerName

    #Draw router layout
    draw_layout(mx_dict)

now = datetime.datetime.now()
fileExcell = ''.join(['/Users/jomih/Documents/20.Juniper/100.Scripts/chassis_to_excel/Layout_', str(now.year), '-', str(now.month), '-', str(now.day), '.xlsx'])
wb.save(fileExcell)
